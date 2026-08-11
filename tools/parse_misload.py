# -*- coding: utf-8 -*-
"""解析 OneDrive\誤裝誤訂 內尚未入庫的 xlsx/xls → 產出 JSON 批次檔
用法： python parse_misload.py [來源資料夾] [輸出資料夾]
狀態檔 .uploaded.json 記錄已入庫的 檔名+mtime，重跑只處理新檔/改過的檔。
"""
import openpyxl, xlrd, sys, json, glob, os, re, warnings, datetime, html as htmlmod
warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8")

# 發送站 → [區號, 區]（貨物追蹤系統 HTML 匯出檔沒有「區」欄，靠這個補）
with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "station_region.json"), encoding="utf-8-sig") as f:
    STATION_REGION = json.load(f)

SRC_DIR = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\26516\OneDrive\誤裝誤訂"
OUT_DIR = sys.argv[2] if len(sys.argv) > 2 else os.path.join(SRC_DIR, ".payloads")
STATE = os.path.join(SRC_DIR, ".uploaded.json")
PASSCODE = "8036"
BATCH = 800
os.makedirs(OUT_DIR, exist_ok=True)

def ymd(v):
    if v is None: return None
    if isinstance(v, datetime.datetime): return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if re.fullmatch(r"\d{8}", s):
        return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"
    return None

def hms(v):
    if v is None: return None
    s = str(v).strip().split(".")[0]
    if re.fullmatch(r"\d{5,6}", s):
        s = s.zfill(6)
        return f"{s[0:2]}:{s[2:4]}:{s[4:6]}"
    return s or None

def txt(v):
    if v is None: return None
    s = str(v).strip().replace("\u3000", " ").strip()
    return s or None

def load_sheets(path):
    """xls/xlsx → {工作表名: 2D list}（日期→datetime、錯誤儲存格→None）"""
    sheets = {}
    if path.lower().endswith(".xls"):
        book = xlrd.open_workbook(path)
        for sh in book.sheets():
            rows = []
            for r in range(sh.nrows):
                row = []
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            row.append(xlrd.xldate.xldate_as_datetime(cell.value, book.datemode))
                        except Exception:
                            row.append(None)
                    elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR):
                        row.append(None)
                    else:
                        row.append(cell.value)
                rows.append(row)
            sheets[sh.name] = rows
    else:
        wb = openpyxl.load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            sheets[ws.title] = [list(r) for r in ws.iter_rows(values_only=True)]
    return sheets

class Sheet:
    """1-indexed cell 存取，相容原本 openpyxl 寫法"""
    def __init__(self, rows):
        self.rows = rows
        self.max_row = len(rows)
        self.max_column = max((len(r) for r in rows), default=0)
    def cell(self, r, c):
        v = None
        if 1 <= r <= len(self.rows) and 1 <= c <= len(self.rows[r - 1]):
            v = self.rows[r - 1][c - 1]
        class _C: value = v
        return _C

def is_html_file(path):
    with open(path, "rb") as f:
        head = f.read(64).lstrip(b"\xef\xbb\xbf \t\r\n")
    return head.startswith(b"<")

def parse_html_file(path):
    """貨物追蹤系統匯出的 HTML 偽 xls：只有明細表（19欄、無區欄），無統計表"""
    with open(path, encoding="utf-8", errors="replace") as f:
        text = f.read()
    details = {}
    for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", text, re.S):
        tds = re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S)
        if len(tds) < 19:
            continue
        c = [htmlmod.unescape(re.sub(r"<[^>]+>", "", x)).replace("\xa0", " ").strip() for x in tds[:19]]
        # 0序號 1責任站 2作業日 3作業時間 4十碼貨號 5發送日期 6發送站 7寄貨人 8收貨地址
        # 9發送件數 10到著站 11註區站 12註區件數 13註區碼 14註區日 15註區時間 16註區人員 17責任站回覆 18商品別
        send_d = ymd(c[5])
        if not send_d or not re.fullmatch(r"\d+", c[0]):
            continue
        send_station = txt(c[6])
        rc, rn = STATION_REGION.get(send_station, (None, None))
        details.setdefault(send_d, []).append({
            "region_code": rc, "region_name": rn,
            "resp_station": txt(c[1]), "seq": int(c[0]),
            "work_date": ymd(c[2]), "work_time": hms(c[3]),
            "tracking_no": txt(c[4]), "send_station": send_station,
            "sender": txt(c[7]), "address": txt(c[8]),
            "pieces": int(float(c[9])) if re.fullmatch(r"\d+(\.\d+)?", c[9]) else None,
            "dest_station": txt(c[10]), "note_station": txt(c[11]),
            "note_code": txt(c[13]), "note_date": ymd(c[14]), "note_time": hms(c[15]),
            "note_staff": txt(c[16]), "reply": txt(c[17]), "product_type": txt(c[18]),
        })
    return [], details

def parse_file(path):
    if is_html_file(path):
        return parse_html_file(path)
    sheets = load_sheets(path)
    if "統計表" not in sheets or "誤裝訂明細" not in sheets:
        raise ValueError("缺少「統計表」或「誤裝訂明細」工作表")
    ws = Sheet(sheets["統計表"])
    date_cols = []
    for c in range(6, ws.max_column + 1):
        v = ws.cell(3, c).value
        if isinstance(v, datetime.datetime):
            date_cols.append((c, v.strftime("%Y-%m-%d")))
    stats = []
    for r in range(5, ws.max_row + 1):
        b, cname, dname = ws.cell(r, 2).value, txt(ws.cell(r, 3).value), txt(ws.cell(r, 4).value)
        if cname is None and dname is None: break
        if cname == "全公司":
            scope, code, region, station = "company", 0, "全公司", ""
        elif dname is None:
            scope, code, region, station = "region", int(b), cname, ""
        elif str(b) == "0" or b == 0:
            scope, code, region, station = "special", 0, "特殊站", dname
        else:
            scope, code, region, station = "station", int(b), cname, dname
        for col, d in date_cols:
            sent = ws.cell(r, col).value
            err = ws.cell(r, col + 1).value
            sent = int(sent) if isinstance(sent, (int, float)) else 0
            err = int(err) if isinstance(err, (int, float)) else 0
            if sent == 0 and err == 0: continue
            stats.append({"stat_date": d, "scope": scope, "region_code": code,
                          "region_name": region, "station_name": station,
                          "sent_count": sent, "err_count": err})
    ws2 = Sheet(sheets["誤裝訂明細"])
    details = {}
    for r in range(5, ws2.max_row + 1):
        seq = ws2.cell(r, 4).value
        send_d = ymd(ws2.cell(r, 9).value)
        if not isinstance(seq, (int, float)) or send_d is None: continue
        row = {
            "region_code": int(ws2.cell(r, 1).value) if isinstance(ws2.cell(r, 1).value, (int, float)) else None,
            "region_name": txt(ws2.cell(r, 3).value),
            "resp_station": txt(ws2.cell(r, 5).value),
            "seq": int(seq),
            "work_date": ymd(ws2.cell(r, 6).value),
            "work_time": hms(ws2.cell(r, 7).value),
            "tracking_no": txt(ws2.cell(r, 8).value),
            "send_station": txt(ws2.cell(r, 10).value),
            "sender": txt(ws2.cell(r, 11).value),
            "address": txt(ws2.cell(r, 12).value),
            "pieces": int(ws2.cell(r, 13).value) if isinstance(ws2.cell(r, 13).value, (int, float)) else None,
            "dest_station": txt(ws2.cell(r, 14).value),
            "note_station": txt(ws2.cell(r, 15).value),
            "note_code": txt(ws2.cell(r, 17).value),
            "note_date": ymd(ws2.cell(r, 18).value),
            "note_time": hms(ws2.cell(r, 19).value),
            "note_staff": txt(ws2.cell(r, 20).value),
            "reply": txt(ws2.cell(r, 21).value),
            "product_type": txt(ws2.cell(r, 22).value),
        }
        details.setdefault(send_d, []).append(row)
    return stats, details

state = {}
if os.path.exists(STATE):
    with open(STATE, encoding="utf-8-sig") as f:
        state = json.load(f)

manifest = []
processed = []
for path in sorted(glob.glob(os.path.join(SRC_DIR, "2026*發送誤裝訂統計表.xls")) + glob.glob(os.path.join(SRC_DIR, "2026*發送誤裝訂統計表.xlsx"))):
    name = os.path.basename(path)
    mtime = int(os.path.getmtime(path))
    if state.get(name) == mtime:
        continue
    m = re.search(r"(\d{8})", name)
    file_date = f"{m.group(1)[0:4]}-{m.group(1)[4:6]}-{m.group(1)[6:8]}"
    stats, details = parse_file(path)
    tag = m.group(1) + ("_xlsx" if path.lower().endswith(".xlsx") else "_xls")
    for i in range(0, len(stats), BATCH):
        fn = os.path.join(OUT_DIR, f"{tag}_stats_{i//BATCH}.json")
        with open(fn, "w", encoding="utf-8") as f:
            json.dump({"passcode": PASSCODE, "action": "upload_stats", "rows": stats[i:i+BATCH]}, f, ensure_ascii=False)
        manifest.append(fn)
    total = 0
    for send_d, rows in details.items():
        total += len(rows)
        for i in range(0, len(rows), BATCH):
            fn = os.path.join(OUT_DIR, f"{tag}_det_{send_d}_{i//BATCH}.json")
            with open(fn, "w", encoding="utf-8") as f:
                json.dump({"passcode": PASSCODE, "action": "upload_details", "send_date": send_d,
                           "first": i == 0, "rows": rows[i:i+BATCH]}, f, ensure_ascii=False)
            manifest.append(fn)
    fn = os.path.join(OUT_DIR, f"{tag}_log.json")
    with open(fn, "w", encoding="utf-8") as f:
        json.dump({"passcode": PASSCODE, "action": "log_upload", "file_date": file_date,
                   "detail_count": total, "stats_days": len(set(s["stat_date"] for s in stats))}, f, ensure_ascii=False)
    manifest.append(fn)
    processed.append((name, mtime))
    print(f"{name}: stats={len(stats)} details={total} dates={sorted(details.keys())}")

with open(os.path.join(OUT_DIR, "manifest.txt"), "w", encoding="utf-8") as f:
    f.write("\n".join(manifest))
with open(os.path.join(OUT_DIR, "pending_state.json"), "w", encoding="utf-8") as f:
    json.dump({"state_file": STATE, "old": state, "new": dict(processed)}, f, ensure_ascii=False)
print(f"new_files={len(processed)} payloads={len(manifest)}")
